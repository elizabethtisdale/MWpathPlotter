import matplotlib
from pandas.tseries import frequencies

matplotlib.use('TkAgg')  #Use the Tkinter backend
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import pyproj
import math
from matplotlib.widgets import Button
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import geopandas as gpd
from shapely.geometry import LineString
import sys
import numpy as np
from matplotlib.patches import Ellipse
import zipfile
import contextily as ctx
from shapely.geometry import LineString, Point
from matplotlib.widgets import Button
import time

# global abriables for dragging detection
is_dragging = [False]
last_mouse_press_pos = [None]
# hover global variables
last_hover_time = 0
hover_interval = 0.5  # 500 ms



# This block sets the necessary environment variables for GDAL and PROJ
# to work correctly when the application is packaged with PyInstaller. (Thanks Gemini)
if getattr(sys, 'frozen', False):
    # If the application is run as a bundle, the PyInstaller bootloader
    # extends the sys module by a flag frozen=True and sets the app path.

    # Add the unpacked C-libraries folder to the system's PATH
    os.environ['PATH'] = sys._MEIPASS + os.pathsep + os.environ['PATH']

    # Set the PROJ and GDAL data paths
    os.environ['PROJ_LIB'] = os.path.join(sys._MEIPASS, 'proj')
    os.environ['GDAL_DATA'] = os.path.join(sys._MEIPASS, 'gdal')


def get_utm_epsg_code(lon):
    """Calculates the appropriate UTM EPSG code for a given longitude."""
    utm_zone = math.floor((lon + 180) / 6) + 1
    # For the US, the northern hemisphere codes (326xx) are used.
    # Build the EPSG code string. "326" is the prefix for WGS 84 / Northern Hemisphere.
    epsg_code = f"326{utm_zone:02d}"
    return epsg_code


def calculate_axis_ranges(lat, lon, distance_miles):
    """
    Calculates 4 new coordinates (N, S, E, W) at a given distance
    from an initial point anywhere in the USA.
    """
    distance_meters = distance_miles * 1609.344
    # Define Coordinate Reference Systems (CRS)
    crs_geographic = pyproj.CRS("EPSG:4269")  # NAD83
    epsg_projected = get_utm_epsg_code(lon)  # to get the code for the UTM zone we are in
    crs_projected = pyproj.CRS(f"EPSG:{epsg_projected}")

    # Objects that will be used to transform the coordinates between systems
    to_utm = pyproj.Transformer.from_crs(crs_geographic, crs_projected, always_xy=True)

    # Project the original coordinate to UTM (meters)
    easting, northing = to_utm.transform(lon, lat)

    # Calculate new points
    points = {}
    points['EAST'] = easting + distance_meters
    points['WEST'] = easting - distance_meters
    points['NORTH'] = northing + distance_meters
    points['SOUTH'] = northing - distance_meters

    # Return UTM ranges and the transformer object
    x_range_utm = (points['WEST'], points['EAST'])
    y_range_utm = (points['SOUTH'], points['NORTH'])
    return x_range_utm, y_range_utm, to_utm, epsg_projected


def get_list_of_paths(df, to_utm):
    """
    Gets the list of start and end points for each path.
    :param df: DataFrame with data read directly from the filtering excel.
    :return: A tuple containing a list of path coordinate tuples and a list of additional info tuples.
    """
    paths = []
    additional_info = []
    frequencies = []

    for i, row in df.iterrows():
        # Transform geographic coordinates to UTM for plotting
        trans_easting, trans_northing = to_utm.transform(row['Lon (NAD83)'], row['Lat (NAD83)'])
        recei_easting, recei_northing = to_utm.transform(row['Lon (NAD83).1'], row['Lat (NAD83).1'])

        transmitter_coordinate = (trans_northing, trans_easting)
        receiver_coordinate = (recei_northing, recei_easting)

        paths.append((transmitter_coordinate, receiver_coordinate))
        additional_info.append((row['Callsign'], row['Path No.']))

        frequencies.append(row['f (MHz)'])

    return paths, additional_info, frequencies

# fresnel zone 2 radius calculation
def fresnel_zone2_radius(freq_mhz, d1, d2):
    # get wavelength
    c = 3e8  # speed of light in m/s
    f_hz = freq_mhz * 1e6  # convert MHz to Hz
    wavelength = c / f_hz

    # calc radius of fresnel zone 2
    radius = np.sqrt((2 * wavelength * d1 * d2) / (d1 + d2))
    return radius

# fresnel zone "line" creator
def fresnel_zone(start_point, end_point, freq_mhz, line_color):
    fresnel_patches = []
    # Calculate the distance and angle between the transmitter and receiver
    dx = end_point[1] - start_point[1]
    dy = end_point[0] - start_point[0]
    path_length = np.sqrt(dx**2 + dy**2)
    angle_deg = np.degrees(np.arctan2(dy, dx))

    num_ellipses = int(path_length / 5)
    num_ellipses = max(50, min(num_ellipses, 1000))

    # Calculate the Fresnel zone radius at each point 
    for n in range(num_ellipses + 1):
        # calc fresnel zone radius at n-th ellipse
        d1 = path_length * (n / num_ellipses)
        d2 = path_length - d1
        r2 = fresnel_zone2_radius(freq_mhz, d1, d2)

        # Position along path
        x_center = start_point[1] + dx * (n / num_ellipses)
        y_center = start_point[0] + dy * (n / num_ellipses)

        # Create ellipse with diameter = 2*r2 for width and height (circle cross-section)
        ellipse = Ellipse((x_center, y_center), width=2*r2, height=2*r2,
                              angle=angle_deg, edgecolor="none", facecolor=line_color, alpha=1)
        fresnel_patches.append(ellipse)

    return fresnel_patches

def export_to_shapefile(full_output_path, paths, attributes, epsg_code):
    """
    Exports a list of microwave paths to a Shapefile using GeoPandas.
    This version does not require arcpy or ArcGIS Pro.
    """
    print(f"Starting GeoPandas export for {len(paths)} selected paths...")

    try:
        attrs_df = pd.DataFrame(attributes, columns=['Callsign', 'Path_Number'])
        # Create LineString geometries from the UTM coordinates
        geometries = [LineString([(path[0][1], path[0][0]), (path[1][1], path[1][0])]) for path in paths]
        gdf = gpd.GeoDataFrame(attrs_df, geometry=geometries, crs=f"EPSG:{epsg_code}")
        gdf['Path_Number'] = gdf['Path_Number'].astype(str)
        gdf.to_file(full_output_path, driver='ESRI Shapefile')

        # --- ZIP all shapefile components ---
        base = os.path.splitext(full_output_path)[0]
        shapefile_dir = os.path.dirname(full_output_path)
        shapefile_files = [f for f in os.listdir(shapefile_dir) if f.startswith(os.path.basename(base))]

        zip_path = base + ".zip"
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for filename in shapefile_files:
                file_path = os.path.join(shapefile_dir, filename)
                zipf.write(file_path, arcname=filename)  # arcname avoids full paths

        # --- Remove loose shapefile files ---
        for filename in shapefile_files:
            os.remove(os.path.join(shapefile_dir, filename))

        print(f"Success! Shapefile saved to: {full_output_path}")
        messagebox.showinfo("Export Successful", f"{len(paths)} paths exported to:\n{full_output_path}")

    except Exception as e:
        print(f"An error occurred during export: {e}")
        messagebox.showerror("Export Error", f"Could not create the Shapefile.\nError: {e}")

def precompute_edge_distances(paths, site_coords, frequencies):
    """
    Precomputes the distance from the WT to the edge of the 2nd Fresnel zone
    for each path and returns a list of those distances in feet.
    """
    from shapely.geometry import LineString, Point

    wt_point = Point(site_coords[0], site_coords[1])
    edge_distances = []

    for i, path in enumerate(paths):
        start_point, end_point = path

        # Create LineString from the path (x = easting, y = northing)
        path_line = LineString([(start_point[1], start_point[0]), (end_point[1], end_point[0])])
        closest_point = path_line.interpolate(path_line.project(wt_point))
        dist_to_path = wt_point.distance(closest_point)

        # Distance from endpoints
        p_start = Point(start_point[1], start_point[0])
        p_end = Point(end_point[1], end_point[0])
        d1 = p_start.distance(closest_point)
        d2 = p_end.distance(closest_point)

        # Compute 2nd Fresnel radius and clearance
        freq_mhz = frequencies[i]
        radius = fresnel_zone2_radius(freq_mhz, d1, d2)
        edge_dist = dist_to_path - radius
        edge_dist_ft = edge_dist * 3.28084

        edge_distances.append(edge_dist_ft)

    return edge_distances

# --- MODIFIED FUNCTION ---
def open_plot_window(parent, paths, x_range, y_range, legend_info, site_coords, radius,
                     input_filepath, epsg_code, frequencies):


    """
    Opens a NEW Tkinter window to display the Matplotlib plot.
    """
    plot_window = tk.Toplevel(parent)
    plot_window.title("Microwave Paths Map")
    plot_window.geometry("1500x1000")

    fig, ax = plt.subplots(figsize=(15, 10))
    fig.subplots_adjust(left=0.1, right=0.75, bottom=0.1, top=0.9)

    # Plot site and radius circle
    radius_meters = radius * 1609.344  # mi to m
    ax.plot(site_coords[0], site_coords[1], 'ro', markersize=10, label='Site Location')
    circle = plt.Circle((site_coords[0], site_coords[1]), radius_meters,
                        color='blue', fill=False, linestyle='--')
    ax.add_artist(circle)
    radius_handle = plt.Line2D([], [], color='blue', linestyle='--', label=f'{radius} mi Radius')

    # Adjust zoom to show just the circle area (with some buffer)
    buffer = radius_meters * 1.1  # 10% padding
    x_center, y_center = site_coords
    home_xlim = (x_center - buffer, x_center + buffer)
    home_ylim = (y_center - buffer, y_center + buffer)

    # Set plot limits to the buffered circle area
    ax.set_xlim(home_xlim)
    ax.set_ylim(home_ylim)

    # Plot each path individually
    all_path_lines = []
    all_fresnel_patches = []
    bright_colors = [
        "#FF0000",  # bright red
        "#00FFFF",  # cyan
        "#FFFF00",  # yellow
        "#006400",   # DarkGreen
        "#FF00FF",  # magenta
        "#FFA500",  # orange
        "#00BFFF",  # deep sky blue
        "#FF1493",  # deep pink
        "#9B30FF",  # electric purple
        "#1E90FF"   # dodger blue
    ]   

    # compute the min distance from the WT to the edge of the 2nd Fresnel zone (for hover tooltip)
    precomputed_edge_dists = precompute_edge_distances(paths, site_coords, frequencies)

    for i, path in enumerate(paths):
        callsign, path_no = legend_info[i]
        start_point, end_point = path
        x_coords = [start_point[1], end_point[1]]
        y_coords = [start_point[0], end_point[0]]

        line_color = bright_colors[i % len(bright_colors)]
        line, = ax.plot(x_coords, y_coords, marker='o', markersize=3, linestyle='-',
                        label=f'{callsign} No.{path_no}', linewidth = .5, color = line_color)

        
        # Store unique ID on the line object for the hover tooltip
        line.set_gid(f"{callsign} No.{path_no}")
        all_path_lines.append(line)

        # --- Fresnel zones ---
        freq_mhz = frequencies[i]  # Frequency in MHz
        fresnel_ellipses = fresnel_zone(start_point, end_point, freq_mhz, line_color)
        
        # add each ellipse to the axes and track it
        for ellipse in fresnel_ellipses:
            ax.add_patch(ellipse)
        all_fresnel_patches.append(fresnel_ellipses)

    legend_map = {}

    # Manually create the legend entries to maintain 1:1 mapping
    custom_handles = [plt.Line2D([], [], color=line.get_color(), label=line.get_label()) for line in all_path_lines]
    custom_handles.insert(0, radius_handle)  # Optional: put radius after site
    custom_handles.insert(0, plt.Line2D([], [], color='red', marker='o', linestyle='', label='Site Location'))

    leg = ax.legend(handles=custom_handles, loc='upper left', bbox_to_anchor=(1.02, 1), title="Paths (Click to toggle)")

    # Now, re-map properly
    path_legend_handles = leg.legend_handles[2:]  # Skip site and radius
    for i, handle in enumerate(path_legend_handles):
        handle.set_picker(5)
        legend_map[handle] = {
            'line': all_path_lines[i],
            'ellipses': all_fresnel_patches[i]
        }


    # Click handler for toggling individual paths via the legend
    def on_pick(event):
        legend_handle = event.artist
        if legend_handle not in legend_map: return

        entry = legend_map[legend_handle]
        line = entry['line']
        ellipses = entry['ellipses']

        new_visibility = not line.get_visible()
        line.set_visible(new_visibility)

        for ellipse in ellipses:
            ellipse.set_visible(new_visibility)

        legend_handle.set_alpha(1.0 if new_visibility else 0.2)
        fig.canvas.draw_idle()

    # --- Hover Tooltip Setup ---
    annot = ax.annotate("", xy=(0, 0), xytext=(10, 10), textcoords="offset points",
                        bbox=dict(boxstyle="round", fc="yellow", ec="black", lw=1))
    annot.set_visible(False)

    # Mouse move handler to show path details on hover
    def on_hover(event):
        global last_hover_time
    
        now = time.time()
        if now - last_hover_time < hover_interval:
            return  # skip processing to throttle

        last_hover_time = now

        if not event.inaxes == ax:
            return

        is_annot_visible = False
        for i, line in enumerate(all_path_lines):
            if line.get_visible() and line.contains(event)[0]:
                label = line.get_gid()
                edge_dist_ft = precomputed_edge_dists[i]
                annot.set_text(f"{label}\nMin dist to Fresnel edge: {edge_dist_ft:.1f} ft")
                annot.xy = (event.xdata, event.ydata)
                annot.set_visible(True)
                is_annot_visible = True
                break

        if not is_annot_visible and annot.get_visible():
            annot.set_visible(False)

        fig.canvas.draw_idle()

    def on_scroll(event):
        base_scale = 1.2
        ax = event.inaxes
        if ax is None:
            return

        # Get current limits
        cur_xlim = ax.get_xlim()
        cur_ylim = ax.get_ylim()

        xdata = event.xdata  # get event x location
        ydata = event.ydata  # get event y location

        if event.button == 'up':
            # zoom in
            scale_factor = 1 / base_scale
        elif event.button == 'down':
            # zoom out
            scale_factor = base_scale
        else:
            scale_factor = 1
            print(event.button)

        new_width = (cur_xlim[1] - cur_xlim[0]) * scale_factor
        new_height = (cur_ylim[1] - cur_ylim[0]) * scale_factor

        relx = (cur_xlim[1] - xdata) / (cur_xlim[1] - cur_xlim[0])
        rely = (cur_ylim[1] - ydata) / (cur_ylim[1] - cur_ylim[0])

        ax.set_xlim([xdata - new_width * (1-relx), xdata + new_width * relx])
        ax.set_ylim([ydata - new_height * (1-rely), ydata + new_height * rely])
        ax.figure.canvas.draw_idle()

        # refresh map view
        if basemap_on[0]:
            add_basemap()

    # --- "Toggle All" Button Setup ---
    button_ax_toggle = fig.add_axes([0.80, 0.15, 0.15, 0.04])  # Define button position
    toggle_button = Button(button_ax_toggle, 'Toggle All', hovercolor='0.9')

    def on_toggle_clicked(event):
        if not all_path_lines: return
        target_visible = not all_path_lines[0].get_visible()
        for i, line in enumerate(all_path_lines):
            line.set_visible(target_visible)
            for ellipse in all_fresnel_patches[i]:
                ellipse.set_visible(target_visible)

        for handle in path_legend_handles:
            handle.set_alpha(1.0 if target_visible else 0.2)
        fig.canvas.draw_idle()

    toggle_button.on_clicked(on_toggle_clicked)
    fig._toggle_button = toggle_button  # Keep a reference to prevent garbage collection

    # --- Basemap ---
    basemap_on = [False]  # mutable toggle state

    def remove_basemap():
        # Safely remove all current tile images
        for img in ax.images[:]:
            try:
                img.remove()
            except Exception:
                pass

    def add_basemap():
        try:
            remove_basemap()  # clear any old basemap tiles
            ctx.add_basemap(ax, crs=f"EPSG:{epsg_code}",
                            source=ctx.providers.Esri.WorldImagery,
                            alpha=0.3)

        except Exception as e:
            print(f"Could not load basemap: {e}")

    def on_basemap_toggle(event):
        basemap_on[0] = not basemap_on[0]
        if basemap_on[0]:
            add_basemap()
        else:
            remove_basemap()
        fig.canvas.draw_idle()



    # Now create button axes and widget
    button_ax = fig.add_axes([0.80, 0.05, 0.15, 0.04])
    basemap_button = Button(button_ax, 'Toggle Basemap', hovercolor='0.9')

    # Connect the toggle function
    basemap_button.on_clicked(on_basemap_toggle)

    # Keep a reference so it doesn't get garbage collected
    fig._basemap_button = basemap_button



    # --- Home Button ---
    button_ax_home = fig.add_axes([0.80, 0.20, 0.15, 0.04])
    home_button = Button(button_ax_home, 'Home', hovercolor='0.9')

    def on_home(event):
        ax.set_xlim(home_xlim)
        ax.set_ylim(home_ylim)

        if basemap_on[0]:
            add_basemap()

        fig.canvas.draw_idle()

    home_button.on_clicked(on_home)
    fig._home_button = home_button

    # --- "Export" Button Setup ---
    button_ax_export = fig.add_axes([0.80, 0.10, 0.15, 0.04])
    export_button = Button(button_ax_export, 'Export Selected to Layer', hovercolor='0.9')

    # To export to an Esri Shapefile
    def on_export_clicked(event):
        visible_paths = []
        visible_attributes = []
        for i, line in enumerate(all_path_lines):
            if line.get_visible():
                visible_paths.append(paths[i])
                visible_attributes.append(legend_info[i])

        if not visible_paths:
            messagebox.showwarning("Warning", "No paths are selected/visible to export.")
            return

        # Save to the same path as the input excel
        input_dir = os.path.dirname(input_filepath)
        output_shp_path = os.path.join(input_dir, "Filtered_MW_paths.shp")
        export_to_shapefile(output_shp_path, visible_paths, visible_attributes, epsg_code)

    export_button.on_clicked(on_export_clicked)
    fig._export_button = export_button  # Keep a reference

    # -- Dragging the map ---
    prev_xlim = [None]
    prev_ylim = [None]
    is_updating_basemap = [False]

    def on_limits_change(event_ax):
        if not basemap_on[0] or is_updating_basemap[0]:
            return

        new_xlim = ax.get_xlim()
        new_ylim = ax.get_ylim()

        if prev_xlim[0] == new_xlim and prev_ylim[0] == new_ylim:
            return

        try:
            is_updating_basemap[0] = True
            prev_xlim[0] = new_xlim
            prev_ylim[0] = new_ylim
            add_basemap()
            fig.canvas.draw_idle()
        finally:
            is_updating_basemap[0] = False


    # --- Final Plot Configuration ---
    ax.set_xlim(x_range)
    ax.set_ylim(y_range)
    ax.set_xlabel("Easting (m)")
    ax.set_ylabel("Northing (m)")
    ax.set_title("Microwave Paths Map (UTM)")
    ax.grid(True)
    ax.set_aspect('equal', adjustable='box')

    # --- Tkinter Integration ---
    canvas = FigureCanvasTkAgg(fig, master=plot_window)
    canvas.draw()
    # Add the Matplotlib toolbar (zoom, save, etc.)
    toolbar = NavigationToolbar2Tk(canvas, plot_window)
    toolbar.update()

    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    toolbar.pan()


    # Connect events to the new canvas
    canvas.mpl_connect('pick_event', on_pick)
    canvas.mpl_connect('motion_notify_event', on_hover)
    canvas.mpl_connect('scroll_event', on_scroll)
    ax.callbacks.connect('xlim_changed', on_limits_change)
    ax.callbacks.connect('ylim_changed', on_limits_change)




    # DO NOT USE plt.show()
    # plt.show()


# --- Main Interface (MAIN) ---

def browse_for_file():
    """Opens a file dialog to select an Excel file."""
    filepath = filedialog.askopenfilename(
        title="Select Filtered MW Excel File",
        filetypes=(("Excel Files", "*.xlsx;*.xlsm"), ("All files", "*.*"))
    )
    if filepath:
        entry_path.delete(0, tk.END)  # Clear the current entry
        entry_path.insert(0, filepath)  # Insert the new path


def start_generation():
    """Gets the path from the entry box and runs the main script logic."""
    filepath = entry_path.get()
    if not filepath:
        messagebox.showwarning("Warning", "Please select an Excel file first.")
        return

    try:
        # Add a try/except block to catch any errors during loading/processing
        mw_path_df = pd.read_excel(filepath, sheet_name='mw_Results', skiprows=1, usecols='F:J, O:Q')
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb['mw_Input']

        site_lat = ws['B2'].value
        site_long = ws['B3'].value
        path_radius = ws['B4'].value  # in miles

        # Obtain plot axis ranges
        x_range, y_range, to_utm, epsg_code = calculate_axis_ranges(site_lat, site_long, path_radius + 1)
        site_coords_utm = to_utm.transform(site_long, site_lat)
        paths, additional_info, frequencies = get_list_of_paths(mw_path_df, to_utm)


        # Call the new function that opens the plot window
        open_plot_window(root, paths, x_range, y_range, additional_info, site_coords_utm, path_radius,
                 input_filepath=filepath, epsg_code=epsg_code, frequencies=frequencies)


    except Exception as e:
        # If something fails, show a clear error message.
        messagebox.showerror("Error", f"An error occurred while processing the file:\n\n{e}")


# --- Tkinter GUI Setup ---
root = tk.Tk()
root.title("Microwave Path Plotter")
root.geometry("500x150")  # Set initial size

def _on_closing():
    """
    To ensure the process is destroyed and it is not running in the backend
    """
    root.quit()  # Detiene el mainloop de tkinter
    root.destroy() # Destruye la ventana y sus widgets
    sys.exit() # Termina el proceso de Python

root.protocol("WM_DELETE_WINDOW", _on_closing)

# Create a frame for better organization
main_frame = tk.Frame(root, padx=10, pady=10)
main_frame.pack(fill=tk.BOTH, expand=True)

# Row 1: Label, Entry, and Browse Button
label_file = tk.Label(main_frame, text="Filtered MW Excel:")
label_file.grid(row=0, column=0, padx=5, pady=10, sticky="w")

entry_path = tk.Entry(main_frame, width=40)
entry_path.grid(row=0, column=1, padx=5, pady=10, sticky="ew")

button_browse = tk.Button(main_frame, text="Browse...", command=browse_for_file)
button_browse.grid(row=0, column=2, padx=5, pady=10)

# Configure the grid to make the entry box expand with the window
main_frame.grid_columnconfigure(1, weight=1)

# Row 2: Generate Button
button_generate = tk.Button(main_frame, text="Generate Plot", command=start_generation)
button_generate.grid(row=1, column=0, columnspan=3, pady=20)

root.mainloop()